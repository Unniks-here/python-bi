from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


def apply_conditional_formats(ws, df: pd.DataFrame, rules: list[dict]) -> None:
    """Apply simple conditional formatting rules to a worksheet."""
    for rule in rules:
        column = rule.get("column")
        if column not in df.columns:
            continue
        col_idx = df.columns.get_loc(column) + 1
        letter = get_column_letter(col_idx)

        gt_val = rule.get("gt")
        gt_fill = rule.get("gt_fill", "C6EFCE")
        gt_font = rule.get("gt_font", "006100")

        lt_val = rule.get("lt")
        lt_fill = rule.get("lt_fill", "FFC7CE")
        lt_font = rule.get("lt_font", "9C0006")

        for cell in ws[letter][1:]:
            try:
                value = float(cell.value)
            except (TypeError, ValueError):
                continue
            if gt_val is not None and value > gt_val:
                cell.fill = PatternFill(start_color=gt_fill, end_color=gt_fill, fill_type="solid")
                cell.font = Font(color=gt_font, bold=True)
            if lt_val is not None and value < lt_val:
                cell.fill = PatternFill(start_color=lt_fill, end_color=lt_fill, fill_type="solid")
                cell.font = Font(color=lt_font, bold=True)


def export_excel(df: pd.DataFrame, path: str, formatting_rules: list[dict] | None = None) -> None:
    """Export DataFrame to a styled Excel file."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.book.active

        # Auto width
        for i, column in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Bold headers
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Enable filters on header row
        ws.auto_filter.ref = ws.dimensions

        if formatting_rules:
            apply_conditional_formats(ws, df, formatting_rules)

